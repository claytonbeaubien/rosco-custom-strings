#!/usr/bin/env python3
"""
build_catalogue.py — Generate daddario_catalogue.json from the D'Addario price list.

Source:  source_data/Daddario Price List - March-2025.xlsx (sheet "Daddario Price List 2025")
Output:  daddario_catalogue.json (project root)

Re-run this any time D'Addario publishes a new price list. Drop the new .xlsx into
source_data/ (or update SOURCE_XLSX below) and run:

    python3 build_catalogue.py

The calculator reads the JSON output, never the xlsx directly.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from datetime import date
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent
SOURCE_XLSX = PROJECT_ROOT / "source_data" / "Daddario Price List - March-2025.xlsx"
SHEET_NAME = "Daddario Price List 2025"
OUTPUT_JSON = PROJECT_ROOT / "daddario_catalogue.json"

MARKUP = {  # mirrored in rosco_string_engine.json -> pricing_config.markup
    "guitar": 1.95,
    "bass":   1.55,
    "acoustic": 1.95,  # reserved; calc doesn't use acoustic yet
}
CURRENCY = "CAD"

# Each prefix maps to series/instrument/type metadata. Order matters only when
# matching codes — we sort by length descending so e.g. NYXLB beats XLB.
PREFIX_MAP = {
    "PL":    {"series": "XL",            "instrument": "guitar",   "type": "plain"},
    "NW":    {"series": "XL",            "instrument": "guitar",   "type": "wound"},
    "NYS":   {"series": "NYXL",          "instrument": "guitar",   "type": "plain"},
    "NYNW":  {"series": "NYXL",          "instrument": "guitar",   "type": "wound"},
    "PSG":   {"series": "ProSteels",     "instrument": "guitar",   "type": "wound"},
    "XLB":   {"series": "XL",            "instrument": "bass",     "type": "wound"},
    "XB":    {"series": "XL",            "instrument": "bass",     "type": "wound"},
    "NYXLB": {"series": "NYXL",          "instrument": "bass",     "type": "wound"},
    "PSB":   {"series": "ProSteels",     "instrument": "bass",     "type": "wound"},
    "NB":    {"series": "Nickel Bronze", "instrument": "acoustic", "type": "wound"},
    "BW":    {"series": "80/20 Bronze",  "instrument": "acoustic", "type": "wound"},
    "XSPL":  {"series": "XS Coated",     "instrument": "guitar",   "type": "plain"},
    "XSNW":  {"series": "XS Coated",     "instrument": "guitar",   "type": "wound"},
}

# Match longest prefixes first so NYXLB wins over XLB, NYNW wins over NW, etc.
PREFIXES_BY_LENGTH = sorted(PREFIX_MAP.keys(), key=len, reverse=True)

# Manual additions — strings sourced outside the standard D'Addario price sheet
# (special-order or distributor extras Clayton actually buys). These are merged
# in after the spreadsheet pass and survive every xlsx re-run. If a code here
# collides with a row in the spreadsheet, the manual entry wins so Clayton's
# real invoiced cost takes precedence over D'Addario's published net.
#
# Format: (prod_code, description, retail_price, net_cost)
# Series/instrument/type/gauge are inferred from the prefix exactly like
# spreadsheet rows.
MANUAL_ADDITIONS = [
    # NW090 — single .090 nickel wound, used as the lowest string in 8-string
    # and extreme drop-tuning guitar packs. Not on D'Addario's standard
    # March 2025 price sheet but available through Clayton's distributor.
    # Source: invoice screenshot, Apr 2026.
    ("NW090", "SINGLE NICKEL WOUND 090", 12.50, 6.13),
    # PL021 — single .021 plain steel. Sits between PL020 and PL022 (which
    # D'Addario does sell as singles). Useful for alternate tunings on
    # heavier-gauged top strings. Stocked.
    # Source: invoice screenshot, Apr 2026.
    ("PL021", "SINGLE PLAIN STEEL 021", 1.80, 0.88),
    # NW065 — single .065 nickel wound. Sits between NW064 and NW066
    # (both of which D'Addario sells as singles). Useful for filling
    # out 6/7-string low-end ladders. Stocked.
    # Source: invoice screenshot, Apr 2026.
    ("NW065", "SINGLE NICKEL WOUND 065", 6.00, 2.94),
]

# Multi-string bundles to exclude (uppercase substring match against description).
EXCLUDE_KEYWORDS = (
    "KIT",
    "SET",
    "3-PACK",
    "5-PACK",
    "10P",
    "3D",
    "2-PACK",
    "PROPACK",
)

DIGIT_RE = re.compile(r"^(\d+)")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def match_prefix(code: str) -> str | None:
    """Return the longest PREFIX_MAP key that `code` starts with, or None."""
    for p in PREFIXES_BY_LENGTH:
        if code.startswith(p):
            return p
    return None


def extract_gauge(code: str, prefix: str) -> tuple[float, str] | None:
    """
    Pull the gauge from a product code.

    The digits after the prefix express the gauge in fixed-point inches:
      "009"  -> .009"      (3 digits, divide by 1_000)
      "0085" -> .0085"     (4 digits half-gauge, divide by 10_000)
      "105"  -> .105"      (3 digits, divide by 1_000)

    `gauge_display` is just "." prepended to the raw digit string, which gives
    the right number of decimals automatically (".009", ".0085", ".105").

    Returns (gauge_float, gauge_display_string) or None if no digits found.
    """
    rest = code[len(prefix):]
    m = DIGIT_RE.match(rest)
    if not m:
        return None
    digits = m.group(1)
    if len(digits) not in (3, 4):
        return None
    gauge = int(digits) / (10 ** len(digits))
    display = "." + digits
    return gauge, display


def is_bundle(description: str) -> bool:
    desc_u = (description or "").upper()
    return any(kw in desc_u for kw in EXCLUDE_KEYWORDS)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def build() -> list[dict]:
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Source spreadsheet not found: {SOURCE_XLSX}")

    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise KeyError(
            f"Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]

    # Locate columns by header so the script survives column reorders.
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {name: header.index(name) for name in
           ("ProdCode", "Description", "Retail Price", "Net Price")
           if name in header}
    missing = {"ProdCode", "Description", "Retail Price", "Net Price"} - col.keys()
    if missing:
        raise KeyError(f"Missing expected columns: {missing}. Got: {header}")

    strings: list[dict] = []
    skipped_bundles = 0
    skipped_no_prefix = 0
    skipped_no_gauge = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[col["ProdCode"]]
        desc = row[col["Description"]]
        retail = row[col["Retail Price"]]
        net = row[col["Net Price"]]

        if not code:
            continue
        code = str(code).strip()

        if is_bundle(desc):
            skipped_bundles += 1
            continue

        prefix = match_prefix(code)
        if prefix is None:
            skipped_no_prefix += 1
            continue

        gauge_info = extract_gauge(code, prefix)
        if gauge_info is None:
            skipped_no_gauge += 1
            continue
        gauge, gauge_display = gauge_info

        meta = PREFIX_MAP[prefix]

        try:
            net_cost = float(net) if net is not None else 0.0
            retail_price = float(retail) if retail is not None else 0.0
        except (TypeError, ValueError):
            continue

        strings.append({
            "prod_code": code,
            "description": (desc or "").strip(),
            "series": meta["series"],
            "instrument": meta["instrument"],
            "type": meta["type"],
            "gauge": gauge,
            "gauge_display": gauge_display,
            "net_cost": net_cost,
            "retail_price": retail_price,
        })

    # Merge in manual additions. These overwrite spreadsheet rows with the
    # same prod_code, since the manual values reflect what Clayton actually
    # pays.
    by_code = {s["prod_code"]: s for s in strings}
    added = 0
    replaced = 0
    for code, desc, retail, net in MANUAL_ADDITIONS:
        prefix = match_prefix(code)
        if prefix is None:
            print(f"  skipping manual addition {code}: no matching prefix")
            continue
        gauge_info = extract_gauge(code, prefix)
        if gauge_info is None:
            print(f"  skipping manual addition {code}: gauge could not be parsed")
            continue
        gauge, gauge_display = gauge_info
        meta = PREFIX_MAP[prefix]
        entry = {
            "prod_code": code,
            "description": desc,
            "series": meta["series"],
            "instrument": meta["instrument"],
            "type": meta["type"],
            "gauge": gauge,
            "gauge_display": gauge_display,
            "net_cost": float(net),
            "retail_price": float(retail),
            "source": "manual",
        }
        if code in by_code:
            replaced += 1
        else:
            added += 1
        by_code[code] = entry
    strings = list(by_code.values())

    # Sort: instrument -> series -> type -> gauge ascending
    strings.sort(key=lambda s: (s["instrument"], s["series"], s["type"], s["gauge"]))

    print(f"Read {ws.max_row - 1} data rows from {SOURCE_XLSX.name}")
    print(f"  skipped {skipped_bundles} bundle/kit rows")
    print(f"  skipped {skipped_no_prefix} rows with non-targeted prefix")
    print(f"  skipped {skipped_no_gauge} rows where gauge could not be parsed")
    print(f"  manual additions: {added} added, {replaced} replaced ({len(MANUAL_ADDITIONS)} total)")
    print(f"  kept    {len(strings)} single strings")
    print()

    return strings


def print_summary(strings: list[dict]) -> None:
    """Print the per-bucket counts that the build instructions require."""

    def count(instrument: str, series: str, type_: str | None = None) -> int:
        return sum(
            1 for s in strings
            if s["instrument"] == instrument
            and s["series"] == series
            and (type_ is None or s["type"] == type_)
        )

    print("Catalogue summary:")
    print(f"  Guitar XL plain:      {count('guitar', 'XL', 'plain'):>3} strings")
    print(f"  Guitar XL wound:      {count('guitar', 'XL', 'wound'):>3} strings")
    print(f"  Guitar NYXL plain:    {count('guitar', 'NYXL', 'plain'):>3} strings")
    print(f"  Guitar NYXL wound:    {count('guitar', 'NYXL', 'wound'):>3} strings")
    print(f"  Guitar ProSteels:     {count('guitar', 'ProSteels'):>3} strings")
    print(f"  Guitar XS plain:      {count('guitar', 'XS Coated', 'plain'):>3} strings")
    print(f"  Guitar XS wound:      {count('guitar', 'XS Coated', 'wound'):>3} strings")
    print(f"  Bass XL:              {count('bass', 'XL'):>3} strings")
    print(f"  Bass NYXL:            {count('bass', 'NYXL'):>3} strings")
    print(f"  Bass ProSteels:       {count('bass', 'ProSteels'):>3} strings")
    print(f"  Acoustic NB:          {count('acoustic', 'Nickel Bronze'):>3} strings")
    print(f"  Acoustic 80/20:       {count('acoustic', '80/20 Bronze'):>3} strings")
    print(f"  TOTAL:                {len(strings):>3} strings")


def main() -> None:
    strings = build()

    output = {
        "meta": {
            "source": SOURCE_XLSX.name,
            "sheet": SHEET_NAME,
            "generated": date.today().isoformat(),
            "currency": CURRENCY,
            "markup": MARKUP,
        },
        "strings": strings,
    }

    OUTPUT_JSON.write_text(json.dumps(output, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {len(strings)} strings -> {OUTPUT_JSON.relative_to(PROJECT_ROOT)}")
    print()
    print_summary(strings)


if __name__ == "__main__":
    main()
